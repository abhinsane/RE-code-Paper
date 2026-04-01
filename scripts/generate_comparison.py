"""
Generate comparison Excel workbook for PQ E-Voting test results.
Previous test: 5 voters, original contract (27-03-2026)
Current test:  30 voters, optimised contract (01-04-2026)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # dark blue  — header rows
C_PREV     = "D6E4F0"   # light blue — previous test
C_CURR     = "D5F5E3"   # light green — current test
C_DIFF     = "FFF3CD"   # yellow     — difference / saving
C_RED      = "FADBD8"   # red        — negative / worse
C_TITLE    = "2E4057"   # title bar

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_white(sz=11):
    return Font(bold=True, color="FFFFFF", size=sz)

def bold(sz=11):
    return Font(bold=True, size=sz)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, values, col_start=1, bg=C_HEADER):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start+i, value=v)
        c.font       = bold_white()
        c.fill       = hfill(bg)
        c.alignment  = center()
        c.border     = thin_border()

def data_row(ws, row, values, col_start=1, bg=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start+i, value=v)
        c.alignment  = center()
        c.border     = thin_border()
        if bg:
            c.fill = hfill(bg)

def title_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", size=13)
    c.fill      = hfill(C_TITLE)
    c.alignment = center()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def pct_change(old, new):
    if old == 0:
        return "N/A"
    return f"{((new - old) / old * 100):+.1f}%"

def saving(old, new):
    if old == 0:
        return "N/A"
    return f"{((old - new) / old * 100):.1f}%"

# ============================================================
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ============================================================
# SHEET 1 — Overview
# ============================================================
ws = wb.create_sheet("Overview")
title_row(ws, 1, "PQ E-Voting — Test Comparison Overview", 4)
header_row(ws, 2, ["Parameter", "Previous Test", "Current Test", "Change"], bg=C_HEADER)

overview = [
    ("Date",                    "27 Mar 2026",          "01 Apr 2026",          "—"),
    ("Contract",                "0xE2B789…90f77",        "0xE7729B…4004F",        "Optimised"),
    ("Network",                 "Sepolia Testnet",       "Sepolia Testnet",       "—"),
    ("Voters registered",       5,                       30,                      "+25"),
    ("Votes cast",              5,                       30,                      "+25"),
    ("Candidates",              4,                       4,                       "—"),
    ("Biometric dataset",       "Synthetic",             "SOCOFing (real)",       "Real data"),
    ("Total transactions",      7,                       33,                      "+26"),
    ("Revocations",             0,                       4,                       "+4"),
    ("Bio failures",            0,                       4,                       "+4"),
    ("Chain length (blocks)",   "N/A",                   2,                       "—"),
    ("Contract version",        "Original (storage)",    "Optimised (events)",    "Gas -63%"),
]

for r, row in enumerate(overview, 3):
    bg = C_CURR if r % 2 == 0 else None
    data_row(ws, r, row, bg=bg)
    if "+" in str(row[3]) or "Real" in str(row[3]) or "Gas" in str(row[3]):
        ws.cell(r, 4).fill = hfill(C_CURR)
    if "N/A" in str(row[1]):
        ws.cell(r, 2).fill = hfill(C_RED)

set_col_widths(ws, [30, 26, 26, 18])

# ============================================================
# SHEET 2 — Gas Cost Comparison
# ============================================================
ws = wb.create_sheet("Gas Costs")
title_row(ws, 1, "Gas Cost Comparison — Previous vs Current Contract", 6)
header_row(ws, 2,
    ["Operation", "Previous Gas", "Current Gas", "Saving", "Previous Cost*", "Current Cost*"],
    bg=C_HEADER)

gas_data = [
    ("anchorVote() per voter",  149895, 56100),    # real: 73,200 first voter, 56,100 typical
    ("recordBatch()",           68879,  68900),    # real: unchanged (array push)
    ("finalizeElection()",      2432679, 154940),  # real: confirmed from Etherscan
    ("Contract deployment",     1500000, 1478856), # real: 1,478,856 gas
]

ETH_PRICE = 3000
GWEI      = 15e-9

for r, (op, prev, curr) in enumerate(gas_data, 3):
    sv   = saving(prev, curr)
    p_cost = f"${prev * GWEI * ETH_PRICE:.2f}"
    c_cost = f"${curr * GWEI * ETH_PRICE:.2f}"
    row  = [op, prev, curr, sv, p_cost, c_cost]
    data_row(ws, r, row)
    ws.cell(r, 2).fill = hfill(C_PREV)
    ws.cell(r, 3).fill = hfill(C_CURR)
    if sv != "N/A" and float(sv.replace("%","")) > 0:
        ws.cell(r, 4).fill = hfill(C_DIFF)
    ws.cell(r, 5).fill = hfill(C_PREV)
    ws.cell(r, 6).fill = hfill(C_CURR)

# totals for 5 voters
r = 7
ws.cell(r, 1, "TOTAL — 5 voters").font = bold()
prev_total = 5*149895 + 68879 + 2432679
curr_total = 5*56100 + 68900 + 154940   # real measured values
ws.cell(r, 2, prev_total).fill = hfill(C_PREV)
ws.cell(r, 3, curr_total).fill = hfill(C_CURR)
ws.cell(r, 4, saving(prev_total, curr_total)).fill = hfill(C_DIFF)
ws.cell(r, 5, f"${prev_total*GWEI*ETH_PRICE:.2f}").fill = hfill(C_PREV)
ws.cell(r, 6, f"${curr_total*GWEI*ETH_PRICE:.2f}").fill = hfill(C_CURR)
for col in range(1,7): ws.cell(r,col).border = thin_border()

# totals for 30 voters (real: 1 cold voter + 29 warm)
r = 8
ws.cell(r, 1, "TOTAL — 30 voters (measured)").font = bold()
prev_30 = 30*149895 + 68879 + 2432679
curr_30 = 73200 + 29*56100 + 68900 + 154940  # real: 73200 first + 56100×29 + fixed
ws.cell(r, 2, prev_30).fill = hfill(C_PREV)
ws.cell(r, 3, curr_30).fill = hfill(C_CURR)
ws.cell(r, 4, saving(prev_30, curr_30)).fill = hfill(C_DIFF)
ws.cell(r, 5, f"${prev_30*GWEI*ETH_PRICE:.2f}").fill = hfill(C_PREV)
ws.cell(r, 6, f"${curr_30*GWEI*ETH_PRICE:.2f}").fill = hfill(C_CURR)
for col in range(1,7): ws.cell(r,col).border = thin_border()

ws.cell(9, 1, "* Cost at 15 gwei / $3,000 ETH (mainnet estimate)").font = Font(italic=True, size=9)

set_col_widths(ws, [28, 16, 16, 12, 16, 16])

# ============================================================
# SHEET 3 — Cost Extrapolation
# ============================================================
ws = wb.create_sheet("Cost Extrapolation")
title_row(ws, 1, "Cost Extrapolation — Optimised Contract (Current)", 7)
header_row(ws, 2,
    ["Voters (n)", "anchorVote gas", "Fixed gas", "Total gas",
     "Cost @ 5 gwei", "Cost @ 15 gwei", "Cost @ 60 gwei"],
    bg=C_HEADER)

scales = [30, 100, 500, 1000, 5000, 10000, 100000, 1000000]
fixed  = 68900 + 154940   # real: recordBatch + finalizeElection

for r, n in enumerate(scales, 3):
    av_gas    = n * 56100   # real measured anchorVote gas (typical)
    total_gas = av_gas + fixed
    c5  = f"${total_gas * 5e-9  * ETH_PRICE:,.2f}"
    c15 = f"${total_gas * 15e-9 * ETH_PRICE:,.2f}"
    c60 = f"${total_gas * 60e-9 * ETH_PRICE:,.2f}"
    bg = C_CURR if r % 2 == 0 else None
    data_row(ws, r, [n, av_gas, fixed, total_gas, c5, c15, c60], bg=bg)

# per-voter row
r = len(scales) + 4
header_row(ws, r, ["Per-voter marginal cost (56,100 gas — real measured)"], bg="4A4A4A")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
data_row(ws, r+1,
    ["1 voter", 56100, 0, 56100,
     f"${56100*5e-9*ETH_PRICE:.4f}",
     f"${56100*15e-9*ETH_PRICE:.4f}",
     f"${56100*60e-9*ETH_PRICE:.4f}"])

set_col_widths(ws, [14, 16, 14, 14, 14, 14, 14])

# ============================================================
# SHEET 4 — Timing Comparison
# ============================================================
ws = wb.create_sheet("Timing")
title_row(ws, 1, "Operation Timing — Previous vs Current Test", 5)
header_row(ws, 2,
    ["Operation", "Previous (5 voters)", "Current (30 voters)", "Unit", "Notes"],
    bg=C_HEADER)

# Previous timing from demo.py output (approximate, not logged in detail)
timing = [
    ("ElectionAuthority init",      "~20,000",  "22,669",   "ms", "PQ keygen + FHE setup"),
    ("EthBridge deploy",            "~1,500",   "1,540",    "ms", "Contract connection"),
    ("register_voter (enroll)",     "~30,000",  "~30,000",  "ms", "Avg per voter"),
    ("authenticate (bio verify)",   "~30,000",  "~42,000",  "ms", "Avg — real SOCOFing slower"),
    ("cast_vote (FHE+ZKP+DSA)",     "~35,000",  "~42,000",  "ms", "Avg per voter"),
    ("receive_vote (ZKP+DSA ver.)", "~30",      "~63",      "ms", "Authority verify"),
    ("eth_bridge.anchor_vote",      "~12,000",  "~10,100",  "ms", "Sepolia tx confirmation"),
    ("process_revocation_request",  "N/A",      "~140,000", "ms", "Bio+DSA verify, 4 done"),
    ("Total per voter (vote)",      "~77,000",  "~94,100",  "ms", "Auth + cast + anchor"),
]

for r, row in enumerate(timing, 3):
    data_row(ws, r, row)
    ws.cell(r, 2).fill = hfill(C_PREV)
    ws.cell(r, 3).fill = hfill(C_CURR)
    if "N/A" in str(row[1]):
        ws.cell(r, 2).fill = hfill(C_RED)

set_col_widths(ws, [32, 22, 22, 8, 38])

# ============================================================
# SHEET 5 — Election Results
# ============================================================
ws = wb.create_sheet("Election Results")
title_row(ws, 1, "Election Results Comparison", 5)
header_row(ws, 2,
    ["Candidate", "Prev Votes", "Prev %", "Curr Votes", "Curr %"],
    bg=C_HEADER)

# Previous test results (from paper docs — 5 voters)
prev_results = [("Modi", 2, 40.0), ("Rahul", 1, 20.0),
                ("Kejriwal", 1, 20.0), ("Mamta", 1, 20.0)]
curr_results = [("Kejriwal", 10, 33.3), ("Rahul", 8, 26.7),
                ("Mamta", 7, 23.3), ("Modi", 5, 16.7)]

# merge by candidate
candidates = ["Modi", "Rahul", "Kejriwal", "Mamta"]
prev_map = {c: (v, p) for c, v, p in prev_results}
curr_map = {c: (v, p) for c, v, p in curr_results}

for r, cand in enumerate(candidates, 3):
    pv, pp = prev_map.get(cand, (0, 0.0))
    cv, cp = curr_map.get(cand, (0, 0.0))
    row = [cand, pv, f"{pp:.1f}%", cv, f"{cp:.1f}%"]
    data_row(ws, r, row)
    ws.cell(r, 2).fill = hfill(C_PREV)
    ws.cell(r, 3).fill = hfill(C_PREV)
    ws.cell(r, 4).fill = hfill(C_CURR)
    ws.cell(r, 5).fill = hfill(C_CURR)

# Winner rows
r = 7
data_row(ws, r, ["Winner", "Modi (40%)", "—", "Kejriwal (33.3%)", "—"])
ws.cell(r, 2).fill = hfill(C_DIFF)
ws.cell(r, 4).fill = hfill(C_DIFF)
for col in range(1,6): ws.cell(r,col).font = bold()

r = 8
data_row(ws, r, ["Total votes", 5, "100%", 30, "100%"])
for col in range(1,6): ws.cell(r,col).font = bold()

set_col_widths(ws, [16, 14, 12, 14, 12])

# ============================================================
# SHEET 6 — Biometric Results
# ============================================================
ws = wb.create_sheet("Biometric")
title_row(ws, 1, "Biometric Evaluation Results", 4)
header_row(ws, 2, ["Metric", "Previous Test", "Current Test", "Notes"], bg=C_HEADER)

bio = [
    ("Dataset",             "Synthetic",        "SOCOFing (real)",      "Real fingerprints"),
    ("Subjects",            "N/A",              "100 (eval)",           "Pre-test evaluation"),
    ("Genuine pairs",       "N/A",              "8,189",                "From bio_eval_results.txt"),
    ("Impostor pairs",      "N/A",              "40,945",               ""),
    ("Genuine mean score",  "1.000",            "0.9089",               "Lower = more realistic"),
    ("Impostor mean score", "N/A",              "0.7461",               ""),
    ("EER",                 "N/A",              "2.61%",                "At threshold 0.818"),
    ("FAR @ threshold",     "N/A",              "8.18%",                "t = 0.818"),
    ("FRR @ threshold",     "N/A",              "0.96%",                "t = 0.818"),
    ("Match threshold",     "0.80",             "0.818",                "EER-optimal"),
    ("Bio failures (live)", "0",                "4",                    "V01 ×1, V30 ×3"),
    ("BioHash score (demo)","1.000 (synthetic)","1.000 (same image)",   "Live capture would ~0.90"),
    ("Revocations tested",  "0",                "4",                    "V21,V26,V29,V31"),
]

for r, row in enumerate(bio, 3):
    data_row(ws, r, row)
    ws.cell(r, 2).fill = hfill(C_PREV)
    ws.cell(r, 3).fill = hfill(C_CURR)
    if "N/A" in str(row[1]):
        ws.cell(r, 2).fill = hfill(C_RED)

set_col_widths(ws, [28, 22, 22, 34])

# ============================================================
# SHEET 7 — Security Events
# ============================================================
ws = wb.create_sheet("Security Events")
title_row(ws, 1, "Security Events — Current Test (30 voters)", 5)
header_row(ws, 2,
    ["Voter", "Event", "Bio Score", "Attempts", "Outcome"],
    bg=C_HEADER)

events = [
    ("V01",  "Bio fail → retry",          "0.000 → 1.000",  2, "Voted ✓"),
    ("V21",  "Biometric revocation",       "—",              1, "Re-enrolled, voted ✓"),
    ("V26",  "Revocation (double-submit)", "—",              2, "Re-enrolled, voted ✓"),
    ("V29",  "Revocation (triple-submit)", "—",              3, "Re-enrolled, voted ✓"),
    ("V30",  "Bio fail ×3 → retry",        "0.000, 0.759, 0.772 → 1.000", 4, "Voted ✓"),
    ("V31",  "Biometric revocation",       "—",              1, "Re-enrolled, voted ✓"),
    ("V20",  "Double-enrol bug",           "—",              2, "Dropped from registry ✗"),
]

colors = [C_CURR, C_CURR, C_DIFF, C_DIFF, C_RED, C_CURR, C_RED]
for r, (row, bg) in enumerate(zip(events, colors), 3):
    data_row(ws, r, row, bg=bg)
    if "✗" in str(row[4]):
        ws.cell(r, 5).fill = hfill(C_RED)

set_col_widths(ws, [8, 28, 32, 10, 24])

# ============================================================
# SHEET 8 — Contract Comparison
# ============================================================
ws = wb.create_sheet("Contract Comparison")
title_row(ws, 1, "Smart Contract — Before vs After Optimisation", 4)
header_row(ws, 2,
    ["Feature", "Original Contract", "Optimised Contract", "Impact"],
    bg=C_HEADER)

contract = [
    ("Contract address",        "0xE2B789…90f77",         "0xE7729B…4004F",       "New deployment"),
    ("anchorVote gas",          "149,895",                "56,100 (meas.)",       "62.6% cheaper"),
    ("finalizeElection gas",    "2,432,679",              "154,940 (meas.)",      "93.6% cheaper"),
    ("Vote evidence storage",   "SSTORE (4 slots/vote)",  "Event only",           "Removed"),
    ("PQ signature storage",    "SSTORE (~103 slots)",    "Event only",           "Removed"),
    ("Nullifier storage",       "SSTORE (1 slot/vote)",   "SSTORE (1 slot/vote)", "Unchanged"),
    ("resultsHash storage",     "SSTORE",                 "SSTORE",               "Unchanged"),
    ("getEvidence() function",  "Present",                "Removed",              "Events only"),
    ("On-chain auditable",      "Yes (storage)",          "Yes (events)",         "Same security"),
    ("Contract-readable",       "Yes",                    "Nullifier + hash only","Sufficient"),
    ("Verified on",             "Routescan",              "Routescan + Sourcify", "Better coverage"),
    ("Test voters",             "5",                      "30",                   "+25"),
    ("Total txs",               "7",                      "33",                   "+26"),
]

for r, row in enumerate(contract, 3):
    data_row(ws, r, row)
    ws.cell(r, 2).fill = hfill(C_PREV)
    ws.cell(r, 3).fill = hfill(C_CURR)
    if "cheaper" in str(row[3]) or "Removed" in str(row[3]):
        ws.cell(r, 4).fill = hfill(C_DIFF)

set_col_widths(ws, [28, 26, 26, 20])

# ============================================================
# Save
# ============================================================
out = "/home/user/RE-code/paper/test-comparison.xlsx"
wb.save(out)
print(f"Saved: {out}")
